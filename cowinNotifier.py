import openpyxl
import requests
from pygame import mixer
from datetime import datetime, timedelta
import time
#https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/

#-------------class to store each session info----------
class sess:
    c_id=""
    c_name=""
    c_address=""
    c_pin=""
    s_id=""
    s_date=""
    s_capacity=""
    s_vaccine=""

#-----------Initialising the final result excel file ---------------
# Workbook is created
wb = openpyxl.load_workbook("final_results.xlsx")
sheet=wb["Sheet"]

def delSheet(sheet):
  
    # continuously delete row 2 untill there
    # is only a single row left over 
    # that contains column names 
    while(sheet.max_row > 1):
        # this method removes the row 2
        sheet.delete_rows(2)
    # return to main function
    return

def putHeader(sheet):
    sheet.cell(row=1,column= 1).value='center id'
    sheet.cell(row=1,column= 2).value= 'center name'
    sheet.cell(row=1,column= 3).value= 'center address'
    sheet.cell(row=1,column= 4).value= 'center pincode'
    sheet.cell(row=1,column= 5).value= 'session id'
    sheet.cell(row=1,column= 6).value= 'date'
    sheet.cell(row=1,column= 7).value= 'Available capacity'
    sheet.cell(row=1,column= 8).value= 'vaccine type'


#----------------The excel file to get state and district input-----------
 
# Give the location of the file
path = "./assets/Districts.xlsx"
 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row

# print (max_row)

# ---------------------Getting states -----------------------

States=[]
test_states=[]
for j in range(2,max_row+1):
    cell_obj_1=sheet_obj.cell(row=j,column=1)
    test_states.append(cell_obj_1.value)

#removing duplicates to get ONLY the state names
[States.append(x) for x in test_states if x not in States]

#print(States)

usr_state=""
index_state=-1
flag=True
while flag:
    print(States)
    print("\nEnter state: ")
    usr_state=input()
    for i in range(0,len(test_states)):
        if usr_state.lower() == test_states[i].lower():
            index_state=i
            flag=False
            break
    if(flag):
        print("No such state found..Enter exact spelling..")

#print(index_state)

#-------------------------getting Districts------------------
usr_dist=""
Districts=[]
index_dist=-1

#storing districts in list for a particular state from excel file
for i in range(index_state+2,max_row+1):
    cell_obj_state=sheet_obj.cell(row=i,column=1)
    if cell_obj_state.value.lower() == usr_state.lower():
        cell_obj_dist=sheet_obj.cell(row=i,column=2)
        Districts.append(cell_obj_dist.value)
        #print(cell_obj_dist.value)
    else:
        break
#print(Districts)

flag=True
while flag:
    print(Districts)
    print("\nEnter district: ")
    usr_dist=input()
    for i in range(0,len(Districts)):
        if usr_dist.lower() == Districts[i].lower():
            index_dist=i
            flag=False
            break
    if(flag):
        print("No such district found..Enter exact spelling..")

#print(index_dist)

#------------getting age, days range and dose no --------------

print("Enter age")
age=int(input())

#print("Enter max days")
#num_days=int(input())

flag=True
dose=1
while flag:
    print("\nEnter dose no(1/2): ")
    dose=int(input())
    if dose==1 or dose ==2:
        flag=False
    if(flag):
        print("Please enter correct dose no (1 or 2)")

print_flag='Y'
dist_id=int(sheet_obj.cell(row=index_state+index_dist+2,column=3).value)
print(dist_id)

#----------------formatting dates to search date wise-------------

date_today=datetime.today()
#list_format=[date_today + timedelta(days=i) for i in range(num_days)]
list_format=[date_today + timedelta(days=i) for i in range(2)]
actual_dates=[i.strftime("%d-%m-%Y")for i in list_format]

ans="y"

#List of all relative sessions
session_list=[sess]

c=0

#----------------starting search for covid vaccine slots------------

while c<2:
    counter=0

    for given_date in actual_dates:

        #URL with the date and district id
        URL = "https://cdn-api.co-vin.in/api/v2/appointment/sessions/public/calendarByDistrict?district_id={}&date={}".format(dist_id, given_date)
        header = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.76 Safari/537.36'}

        #sending http request
        result=requests.get(URL, headers=header)
        if result.ok:
            #creating a json file to extract data in a simpler way
            response_json=result.json()
            if response_json["centers"]:
                if(print_flag.lower()=='y'):
                    print("for date ={}".format(given_date))
                    for center in response_json["centers"]:
                        for session in center["sessions"]:
                            if session["min_age_limit"]<=age and session["available_capacity"]>0 :
                                #class object
                                s=sess()
                                s.c_id=center["center_id"]
                                s.c_name=center["name"]
                                s.c_address=center["address"]
                                s.c_pin=center["pincode"]
                                s.s_id=session["session_id"]
                                s.s_date=session["date"]
                                s.s_capacity=session["available_capacity"]
                                if session["vaccine"] != "":
                                    s.s_vaccine=session["vaccine"]                               
                                session_list.append(s)
        else:
            print("no response")
        c=2

#sorting the slots in lexicographic orders of the center
session_list.sort(key=lambda x: x.c_name)
delSheet(sheet)
putHeader(sheet)
for i in range(len(session_list)):
    j=i+2
    sheet.cell(row=j,column=1,).value=session_list[i].c_id
    sheet.cell(row=j,column=2,).value=session_list[i].c_name
    sheet.cell(row=j,column=3,).value=session_list[i].c_address
    sheet.cell(row=j,column=4,).value=session_list[i].c_pin
    sheet.cell(row=j,column=5,).value=session_list[i].s_id
    sheet.cell(row=j,column=6,).value=session_list[i].s_date
    sheet.cell(row=j,column=7,).value=session_list[i].s_capacity
    sheet.cell(row=j,column=8,).value=session_list[i].s_vaccine

wb.save("final_results.xlsx")    
    
for x in session_list:
    print(x.c_id)
    print(x.c_name)
    print(x.s_id)
    print(x.s_date)
    print("\n")